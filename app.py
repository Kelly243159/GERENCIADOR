from fasthtml.common import *  # FastHTML + Starlette + HTMX helpers
import pandas as pd
from datetime import datetime
from unicodedata import normalize
import re, io, os, smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from starlette.responses import StreamingResponse

# -------------------------------------------------
# Configurações de E-mail (Modificável via variáveis de ambiente)
# -------------------------------------------------
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
EMAIL_USER = os.getenv("EMAIL_USER", "certificadosmvcontabilidade@gmail.com")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "zrfs pbqm urcr viyx")
USE_TLS = os.getenv("USE_TLS", "True").lower() == "true"


# -------------------------------------------------
# Helpers
# -------------------------------------------------
def _only_digits(s: str) -> str:
    return re.sub(r"\D+", "", str(s) if s is not None else "")


def _norm(s: str) -> str:
    s = "" if s is None else str(s)
    s = normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    return re.sub(r"[^a-z0-9 ]+", "", s.strip().lower())


def _pick_col(df, candidates):
    m = {_norm(c): c for c in df.columns}
    for c in candidates:
        k = _norm(c)
        if k in m:
            return m[k]
    return None


def _status(venc_dt, today=None):
    if today is None:
        today = datetime.today()
    if pd.isna(venc_dt):
        return "Sem data"
    d = (venc_dt - today).days
    if d < 0:
        return "Vencido"
    if d <= 30:
        return "A vencer"
    return "No prazo"


def gerar_relatorio(df_sieg, df_cert):
    cnpj_sieg = _pick_col(df_sieg, ["cpf_cnpj", "cnpj", "cpf cnpj", "cpf/cnpj"])
    cnpj_cert = _pick_col(df_cert, ["cpf_cnpj", "cnpj", "cpf cnpj", "cpf/cnpj"])
    if not cnpj_sieg or not cnpj_cert:
        raise ValueError("Não encontrei a coluna CPF_CNPJ/CNPJ em uma das planilhas.")

    col_resp = _pick_col(df_sieg, ["responsavel", "responsável"])
    col_emp = _pick_col(df_sieg, ["empresa", "razao social", "razão social", "cliente", "nome do cliente"])
    col_email = _pick_col(df_sieg, ["email", "e-mail", "email1", "e-mail1"])
    col_venc = _pick_col(df_cert, ["vencimento", "validade", "data de vencimento", "data vencimento"])

    df_sieg["_CPF_CNPJ_"] = df_sieg[cnpj_sieg].map(_only_digits)
    df_cert["_CPF_CNPJ_"] = df_cert[cnpj_cert].map(_only_digits)

    keep = ["_CPF_CNPJ_"] + ([col_venc] if col_venc else [])
    df_cert_small = df_cert[keep].copy()

    merged = pd.merge(df_sieg, df_cert_small, on="_CPF_CNPJ_", how="left")

    out = pd.DataFrame()
    out["Responsavel"] = merged[col_resp] if col_resp else ""
    out["Empresa"] = merged[col_emp] if col_emp else ""
    out["Email"] = merged[col_email] if col_email else ""
    out["CPF_CNPJ"] = merged["_CPF_CNPJ_"]

    if col_venc:
        venc = pd.to_datetime(merged[col_venc], errors="coerce", dayfirst=True)
        out["Vencimento"] = venc.dt.strftime("%d/%m/%Y").fillna("")
        out["Status"] = [_status(d) for d in venc]
    else:
        out["Vencimento"] = ""
        out["Status"] = "Sem data"

    return out


def make_excel_bytes(df: pd.DataFrame, sheet_name="Relatorio") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Cores para os status
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # Aplicar estilo ao cabeçalho e ajustar largura
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = max(12, len(str(df.columns[col - 1])) + 4)

        # Aplicar cores às linhas baseadas no status
        status_col_idx = df.columns.get_loc("Status") + 1  # +1 porque Excel começa em 1

        for row in range(2, len(df) + 2):  # Começa na linha 2 (após cabeçalho)
            status_cell = ws.cell(row=row, column=status_col_idx)
            status_value = status_cell.value

            fill_color = gray_fill  # padrão para "Sem data"
            if status_value == "Vencido":
                fill_color = red_fill
            elif status_value == "A vencer":
                fill_color = yellow_fill
            elif status_value == "No prazo":
                fill_color = green_fill

            # Aplicar cor a todas as células da linha
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill = fill_color
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")

    buf.seek(0)
    return buf.getvalue()


# -------------------------------------------------
# E-mail via SMTP (Funciona em qualquer ambiente)
# -------------------------------------------------
def enviar_email_smtp(destinatario: str, assunto: str, body_html: str):
    """
    Envia e-mail via SMTP - funciona em qualquer ambiente (Windows/Linux/Web)
    """
    try:
        if not all([SMTP_SERVER, EMAIL_USER, EMAIL_PASSWORD]):
            return False, "Configuração de e-mail incompleta. Configure SMTP_SERVER, EMAIL_USER e EMAIL_PASSWORD."

        # Criar mensagem
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = destinatario
        msg['Subject'] = assunto

        # Adicionar corpo HTML
        msg.attach(MIMEText(body_html, 'html'))

        # Conectar e enviar
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            if USE_TLS:
                server.starttls()
            server.login(EMAIL_USER, EMAIL_PASSWORD)
            server.send_message(msg)

        return True, "E-mail enviado com sucesso"
    except Exception as e:
        return False, f"Falha ao enviar e-mail: {str(e)}"


def corpo_email_vencido(empresa: str, vencimento: str) -> str:
    return f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#111">
      <p>Prezado(a) Cliente,</p>
      <p>Identificamos que o certificado digital da empresa <strong>{empresa}</strong> 
      encontra-se <strong style='color:#b91c1c'>VENCIDO</strong> desde <strong>{vencimento}</strong>.</p>
      <p style="margin-top:18px"><strong>⚠️ É urgente regularizar</strong> para evitar:
      <ul>
        <li>Bloqueio de acesso a sistemas governamentais (Receita, SEFAZ, Prefeituras)</li>
        <li>Impossibilidade de emissão de NF-e/NFS-e</li>
        <li>Interrupção de procurações e transmissões</li>
        <li>Multas e outras penalidades</li>
      </ul>
      </p>
      <div style="margin:18px 0;padding:14px 16px;border:1px solid #e5e7eb;border-radius:10px;background:#f8fafc">
        <p style="margin:0 0 6px 0"><strong>💠 MV CONTABILIDADE | CERTIFICADORA DIGITAL</strong></p>
        <ul style="margin:8px 0 0 18px">
          <li>Emissão/renovação de certificados <strong>A1</strong> e <strong>A3</strong></li>
          <li>Atendimento rápido e acompanhamento completo</li>
          <li>Suporte técnico para instalação e uso</li>
          <li>Condições especiais para clientes MV</li>
        </ul>
      </div>
      <p>Ficamos à disposição para proceder com a renovação imediatamente.</p>
      <p>
        📞 <strong>Contato MV:</strong> (41) 99673-1918<br>
        ✉️ <strong>E-mail:</strong> mayke@mvcontabilidade.com.br
      </p>
      <p>Atenciosamente,<br>
      <strong>Equipe MV Contabilidade</strong></p>
    </div>
    """


def corpo_email_a_vencer(empresa: str, vencimento: str) -> str:
    return f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#111">
      <p>Prezado(a) Cliente,</p>
      <p>Identificamos que o certificado digital da empresa <strong>{empresa}</strong> 
      <strong style='color:#f59e0b'>VENCE EM BREVE</strong> no dia <strong>{vencimento}</strong>.</p>
      <p style="margin-top:18px"><strong>📅 É importante renovar antecipadamente</strong> para evitar:
      <ul>
        <li>Bloqueio de acesso a sistemas governamentais</li>
        <li>Impossibilidade de emissão de documentos fiscais</li>
        <li>Interrupção das atividades da empresa</li>
      </ul>
      </p>
      <div style="margin:18px 0;padding:14px 16px;border:1px solid #e5e7eb;border-radius:10px;background:#f8fafc">
        <p style="margin:0 0 6px 0"><strong>💠 MV CONTABILIDADE | CERTIFICADORA DIGITAL</strong></p>
        <ul style="margin:8px 0 0 18px">
          <li>Emissão/renovação de certificados <strong>A1</strong> e <strong>A3</strong></li>
          <li>Atendimento rápido e acompanhamento completo</li>
          <li>Suporte técnico para instalação e uso</li>
          <li>Condições especiais para clientes MV</li>
        </ul>
      </div>
      <p>Entre em contato conosco para renovar seu certificado.</p>
      <p>
        📞 <strong>Contato MV:</strong> (41) 99673-1918<br>
        ✉️ <strong>E-mail:</strong> mayke@mvcontabilidade.com.br
      </p>
      <p>Atenciosamente,<br>
      <strong>Equipe MV Contabilidade</strong></p>
    </div>
    """


# -------------------------------------------------
# FastHTML app
# -------------------------------------------------
app, rt = fast_app()
current_data = {}


def global_css():
    return Style(
        """
        :root { --bg:#0f172a; --panel: rgba(255,255,255,.06); --border: rgba(255,255,255,.15); --text:#e5f2ff; --muted:#94a3b8; --primary:#6366f1; --danger:#ef4444; --warn:#f59e0b; --ok:#10b981; }
        *{box-sizing:border-box}
        body{margin:0;background:linear-gradient(135deg,#0b1020,#0f172a);color:var(--text);font-family:Segoe UI,system-ui,-apple-system,sans-serif}
        .container{max-width:1100px;margin:0 auto;padding:24px}
        .glass{background:var(--panel);backdrop-filter:blur(16px);border:1px solid var(--border);border-radius:18px;padding:28px;margin-bottom:24px}
        .lbl{font-weight:700;margin-bottom:8px;display:block}
        .filebox{width:100%;padding:14px;border:2px dashed var(--primary);border-radius:12px;background:rgba(255,255,255,.08);color:var(--text)}
        .btn{background:linear-gradient(135deg,var(--primary),#4f46e5);color:#fff;border:0;border-radius:12px;padding:14px 22px;font-weight:700;cursor:pointer}
        .grid{display:grid;gap:14px;grid-template-columns:repeat(auto-fit,minmax(220px,1fr))}
        .card{border:1px solid var(--border);background:rgba(255,255,255,.05);border-radius:14px;padding:18px}
        .k{font-size:13px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}
        .v{font-size:28px;font-weight:800;margin-top:6px}
        .danger{color:#fecaca}
        .warn{color:#fde68a}
        .ok{color:#bbf7d0}
        .muted{color:var(--muted)}
        table{width:100%;border-collapse:collapse}
        th,td{padding:12px;border-bottom:1px solid rgba(255,255,255,.08);font-size:14px}
        th{text-align:left;background:rgba(255,255,255,.08)}
        .badge{display:inline-block;padding:6px 10px;border-radius:16px;font-size:12px;font-weight:700;border:1px solid rgba(255,255,255,.2)}
        .bd-danger{background:rgba(239,68,68,.15);color:#fca5a5;border-color:rgba(239,68,68,.3)}
        .bd-warn{background:rgba(245,158,11,.15);color:#fcd34d;border-color:rgba(245,158,11,.3)}
        .bd-ok{background:rgba(16,185,129,.15);color:#6ee7b7;border-color:rgba(16,185,129,.3)}
        .bd-muted{background:rgba(148,163,184,.15);color:#cbd5e1;border-color:rgba(148,163,184,.3)}
        .center{text-align:center}
        .mt-2{margin-top:12px}
        .mt-3{margin-top:18px}
        .mt-4{margin-top:24px}
        .cert-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:16px;margin-top:20px}
        .cert-card{border:1px solid;border-radius:12px;padding:16px;background:rgba(255,255,255,.05)}
        .cert-card.danger{border-color:#ef4444;background:rgba(239,68,68,.08)}
        .cert-card.warn{border-color:#f59e0b;background:rgba(245,158,11,.08)}
        .cert-card.ok{border-color:#10b981;background:rgba(16,185,129,.08)}
        .cert-card.muted{border-color:#94a3b8;background:rgba(148,163,184,.08)}
        .cert-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px}
        .cert-title{font-weight:700;font-size:16px}
        .cert-detail{margin:4px 0;font-size:14px}
        .cert-email{color:var(--muted);font-size:13px;margin-top:8px}
        .loading{{opacity:0.6;pointer-events:none}}
        """
    )


def page():
    return (
        global_css(),
        Main(
            Section(
                Article(
                    Div(
                        Div("🌐", style="font-size:64px;text-shadow:0 8px 24px rgba(0,0,0,.35);text-align:center"),
                        H1("SIEG x Certificados", style="text-align:center;margin:6px 0 4px"),
                        P("Sistema inteligente para gerenciamento e notificação de certificados digitais",
                          style="text-align:center;color:var(--muted)"),
                    ),
                    Form(
                        Div(
                            Label("📋 Planilha SIEG (Clientes e Responsáveis)", cls="lbl"),
                            Input(type="file", name="file_sieg", accept=".xlsx,.xls", required=True, cls="filebox")
                        ),
                        Div(
                            Label("📑 Planilha de Certificados (Datas de Vencimento)", cls="lbl",
                                  style="margin-top:12px"),
                            Input(type="file", name="file_cert", accept=".xlsx,.xls", required=True, cls="filebox")
                        ),
                        Div(Button("🚀 Processar Planilhas", type="submit", cls="btn",
                                   style="margin-top:18px;width:100%")),
                        method="post", action="/processar-upload", enctype="multipart/form-data"
                    ),
                    Div(id="resultados"),
                    cls="glass"
                ),
                cls="container"
            ),
        )
    )


@app.get("/")
def index():
    return Titled("SIEG x Certificados - Dashboard", page())


@app.post("/processar-upload")
async def processar_upload(request):
    try:
        form = await request.form()
        file_sieg = form.get("file_sieg")
        file_cert = form.get("file_cert")
        if not file_sieg or not file_cert:
            return Div(Div(P("❌ Por favor, selecione ambas as planilhas."), cls="glass"), cls="container")

        df_sieg = pd.read_excel(io.BytesIO(await file_sieg.read()), dtype=str)
        df_cert = pd.read_excel(io.BytesIO(await file_cert.read()), dtype=str)

        df = gerar_relatorio(df_sieg, df_cert)

        # Guarda em memória (para ações de e-mail)
        import hashlib
        data_id = hashlib.md5(str(datetime.now()).encode()).hexdigest()[:8]
        current_data[data_id] = df

        # Bytes do Excel (opcional para baixar se quiser)
        xbytes = make_excel_bytes(df)
        current_data[f"{data_id}_excel"] = xbytes

        # Resumo simples
        total = len(df)
        vencidos = (df['Status'] == 'Vencido').sum()
        avencer = (df['Status'] == 'A vencer').sum()
        noprazo = (df['Status'] == 'No prazo').sum()
        semdata = (df['Status'] == 'Sem data').sum()

        # Visualização em cards para certificados vencidos
        cards_vencidos = []
        if vencidos > 0:
            for idx, row in df[df['Status'] == 'Vencido'].iterrows():
                email = (row.get('Email') or '').strip()
                empresa = row.get('Empresa', '') or 'Não informado'
                vencimento = row.get('Vencimento', '') or 'Não informado'

                # Usar índice como identificador único
                card_id = f"vencido_{data_id}_{idx}"

                card_content = [
                    Div(
                        Span("🚨 VENCIDO", cls="badge bd-danger"),
                        Button("📤 Enviar E-mail", cls="btn",
                               style="padding:6px 12px;font-size:12px",
                               hx_post=f"/envia-email-individual/{data_id}/vencido/{idx}",
                               hx_target=f"#res_{card_id}",
                               hx_swap="innerHTML",
                               id=f"btn_{card_id}") if email and '@' in email else
                        Span("E-mail inválido", style="color:var(--muted);font-size:12px"),
                        cls="cert-header"
                    ),
                    Div(Strong(empresa), cls="cert-title"),
                    Div(f"Vencimento: {vencimento}", cls="cert-detail"),
                    Div(f"CNPJ: {row.get('CPF_CNPJ', '')}", cls="cert-detail"),
                ]

                if email and '@' in email:
                    card_content.append(Div(f"📧 {email}", cls="cert-email"))

                card_content.append(Div(id=f"res_{card_id}"))

                cards_vencidos.append(
                    Div(*card_content, cls="cert-card danger")
                )

        # Visualização em cards para certificados a vencer
        cards_a_vencer = []
        if avencer > 0:
            for idx, row in df[df['Status'] == 'A vencer'].iterrows():
                email = (row.get('Email') or '').strip()
                empresa = row.get('Empresa', '') or 'Não informado'
                vencimento = row.get('Vencimento', '') or 'Não informado'

                # Usar índice como identificador único
                card_id = f"avencer_{data_id}_{idx}"

                card_content = [
                    Div(
                        Span("⚠️ A VENCER", cls="badge bd-warn"),
                        Button("📤 Enviar E-mail", cls="btn",
                               style="padding:6px 12px;font-size:12px",
                               hx_post=f"/envia-email-individual/{data_id}/avencer/{idx}",
                               hx_target=f"#res_{card_id}",
                               hx_swap="innerHTML",
                               id=f"btn_{card_id}") if email and '@' in email else
                        Span("E-mail inválido", style="color:var(--muted);font-size:12px"),
                        cls="cert-header"
                    ),
                    Div(Strong(empresa), cls="cert-title"),
                    Div(f"Vencimento: {vencimento}", cls="cert-detail"),
                    Div(f"CNPJ: {row.get('CPF_CNPJ', '')}", cls="cert-detail"),
                ]

                if email and '@' in email:
                    card_content.append(Div(f"📧 {email}", cls="cert-email"))

                card_content.append(Div(id=f"res_{card_id}"))

                cards_a_vencer.append(
                    Div(*card_content, cls="cert-card warn")
                )

        return Div(
            Div(
                H2("📊 Resumo", style="margin:0 0 10px"),
                Div(
                    Div(Div(Span("Total", cls="k"), Div(str(total), cls="v")), cls="card"),
                    Div(Div(Span("Vencidos", cls="k"), Div(str(vencidos), cls="v danger")), cls="card"),
                    Div(Div(Span("A vencer (30 dias)", cls="k"), Div(str(avencer), cls="v warn")), cls="card"),
                    Div(Div(Span("No prazo", cls="k"), Div(str(noprazo), cls="v ok")), cls="card"),
                    Div(Div(Span("Sem data", cls="k"), Div(str(semdata), cls="v muted")), cls="card"),
                    cls="grid"
                ),
                cls="glass"
            ),
            Div(
                H3("🚨 Certificados Vencidos"),
                P("Envie os avisos diretamente via SMTP."),
                Div(
                    Button("📤 Enviar E-mails para Todos os Vencidos", cls="btn",
                           hx_post=f"/envia-emails/{data_id}/vencido",
                           hx_target="#resultado_envio_vencido",
                           hx_indicator="#loading_vencido",
                           style="background:linear-gradient(135deg,#ef4444,#dc2626)"),
                    Span("⏳", id="loading_vencido", style="display:none;margin-left:10px"),
                    A("⬇️ Baixar Excel Colorido", href=f"/baixar-excel/{data_id}", cls="btn",
                      style="margin-left:10px;background:linear-gradient(135deg,#10b981,#059669)"),
                    cls="mt-2"
                ),
                Div(
                    *cards_vencidos if cards_vencidos else
                    [Div(P("✅ Nenhum certificado vencido encontrado."), cls="center")],
                    cls="cert-grid"
                ),
                Div(id="resultado_envio_vencido", cls="mt-3"),
                cls="glass"
            ),
            Div(
                H3("⚠️ Certificados a Vencer (30 dias)"),
                P("Avisos preventivos para certificados que vencerão em breve."),
                Div(
                    Button("📤 Enviar E-mails para Todos a Vencer", cls="btn",
                           hx_post=f"/envia-emails/{data_id}/avencer",
                           hx_target="#resultado_envio_avencer",
                           hx_indicator="#loading_avencer",
                           style="background:linear-gradient(135deg,#f59e0b,#d97706)"),
                    Span("⏳", id="loading_avencer", style="display:none;margin-left:10px"),
                    cls="mt-2"
                ),
                Div(
                    *cards_a_vencer if cards_a_vencer else
                    [Div(P("✅ Nenhum certificado a vencer encontrado."), cls="center")],
                    cls="cert-grid"
                ),
                Div(id="resultado_envio_avencer", cls="mt-3"),
                cls="glass"
            ),
            cls="container"
        )

    except Exception as e:
        import traceback
        print("Erro detalhado:\n", traceback.format_exc())
        return Div(
            Div(H3("❌ Erro ao processar as planilhas"), P(str(e)),
                P("Verifique as colunas (CPF/CNPJ, Vencimento, Email, etc.)."), cls="glass"),
            cls="container"
        )


@app.get("/baixar-excel/{data_id}")
def baixar_excel(request, data_id: str):
    if f"{data_id}_excel" not in current_data:
        return Titled("Erro", Main(Section(Article(P("❌ Dados não encontrados"), cls="glass"), cls="container")))
    xbytes = current_data[f"{data_id}_excel"]
    fname = f"Relatorio_Certificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"', "X-Content-Type-Options": "nosniff"}
    return StreamingResponse(io.BytesIO(xbytes),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)


@app.post("/envia-emails/{data_id}/{tipo}")
async def envia_emails_lote(request, data_id: str, tipo: str):
    df = current_data.get(data_id)
    if df is None:
        return Div(P("❌ Dados não encontrados"), cls="glass")

    if tipo == "vencido":
        filtro = df['Status'] == 'Vencido'
        assunto_base = "AVISO: Certificado Digital Vencido - "
        funcao_corpo = corpo_email_vencido
        target_id = "resultado_envio_vencido"
    elif tipo == "avencer":
        filtro = df['Status'] == 'A vencer'
        assunto_base = "ALERTA: Certificado Digital Vence em Breve - "
        funcao_corpo = corpo_email_a_vencer
        target_id = "resultado_envio_avencer"
    else:
        return Div(P("❌ Tipo de e-mail inválido"), cls="glass")

    certificados = df[filtro]
    if certificados.empty:
        return Div(P(f"✅ Nenhum certificado {tipo.replace('_', ' ')} encontrado."), cls="glass", id=target_id)

    enviados, erros = 0, 0
    linhas = []

    for idx, cert in certificados.iterrows():
        email = (cert.get('Email') or '').strip()
        if not email or '@' not in email:
            continue

        empresa = cert.get('Empresa', 'Cliente') or 'Cliente'
        vencimento = cert.get('Vencimento', 'Data não informada') or 'Data não informada'

        assunto = f"{assunto_base}{empresa}"
        html = funcao_corpo(empresa, vencimento)

        ok, msg = enviar_email_smtp(email, assunto, html)

        if ok:
            enviados += 1
            linhas.append(
                Tr(Td("✅"), Td(empresa), Td(vencimento), Td(email), Td("Enviado"))
            )
        else:
            erros += 1
            linhas.append(
                Tr(Td("❌"), Td(empresa), Td(vencimento), Td(email), Td(msg))
            )

    resumo = Div(
        P(Strong("Resumo: "), Span(f"{enviados} enviados • {erros} erros • {len(certificados)} total", cls="muted")),
        cls="mt-2"
    )

    tabela = Table(
        Thead(Tr(Th("Status"), Th("Empresa"), Th("Vencimento"), Th("E-mail"), Th("Mensagem"))),
        Tbody(*linhas)
    ) if linhas else P("Nenhum e-mail válido para enviar.")

    return Div(
        resumo,
        tabela,
        cls="glass",
        id=target_id
    )


@app.post("/envia-email-individual/{data_id}/{tipo}/{idx}")
def envia_email_individual(request, data_id: str, tipo: str, idx: str):
    df = current_data.get(data_id)
    if df is None:
        return Span("❌ Dados não encontrados", style="color:#ef4444;font-weight:600")

    try:
        idx = int(idx)
        cert = df.iloc[idx]
    except (ValueError, IndexError):
        return Span("❌ Certificado não encontrado", style="color:#ef4444;font-weight:600")

    if tipo == "vencido":
        assunto_base = "AVISO: Certificado Digital Vencido - "
        funcao_corpo = corpo_email_vencido
    elif tipo == "avencer":
        assunto_base = "ALERTA: Certificado Digital Vence em Breve - "
        funcao_corpo = corpo_email_a_vencer
    else:
        return Span("❌ Tipo de e-mail inválido", style="color:#ef4444;font-weight:600")

    email = (cert.get('Email') or '').strip()
    if not email or '@' not in email:
        return Span("❌ E-mail inválido", style="color:#ef4444;font-weight:600")

    empresa = cert.get('Empresa', 'Cliente') or 'Cliente'
    vencimento = cert.get('Vencimento', 'Data não informada') or 'Data não informada'

    assunto = f"{assunto_base}{empresa}"
    html = funcao_corpo(empresa, vencimento)

    ok, msg = enviar_email_smtp(email, assunto, html)

    if ok:
        return Span("✅ E-mail enviado!", style="color:#10b981;font-weight:600")
    else:
        return Span(f"❌ Erro: {msg}", style="color:#ef4444;font-weight:600")


if __name__ == "__main__":
    # Configurações para desenvolvimento
    if not EMAIL_USER:
        print("⚠️  AVISO: Configure as variáveis de ambiente para e-mail:")
        print("   - EMAIL_USER: Seu e-mail")
        print("   - EMAIL_PASSWORD: Sua senha/app password")
        print("   - SMTP_SERVER: servidor SMTP (padrão: smtp.gmail.com)")
        print("   - SMTP_PORT: porta SMTP (padrão: 587)")

    # Rode: uvicorn nome_do_arquivo:app --reload --port 5001
    serve(port=5001)